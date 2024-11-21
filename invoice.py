import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import unicodedata
from Levenshtein import distance
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import difflib
from datetime import datetime  # 日付を取得するためのモジュールをインポート


def get_close_match_with_score(substring, full_string):
    best_match = difflib.get_close_matches(substring, [full_string], n=1, cutoff=0.0)
    if best_match:
        similarity_score = difflib.SequenceMatcher(None, best_match[0], substring).ratio()
        return best_match[0], similarity_score
    else:
        return None, 0


def translate_to_hankaku(text):
    hankaku = (
        "0123456789" "abcdefghijklmnopqrstuvwxyz" "ABCDEFGHIJKLMNOPQRSTUVWXYZ" "-!\"#$%&'()*+,-./:;<=>?@[\\]^_`{|}~"
    )
    zenkaku = (
        "０１２３４５６７８９"
        "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
        "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
        "－！＂＃＄％＆＇（）＊＋，－．／：；＜＝＞？＠［＼］＾＿｀｛｜｝～"
    )
    zenkaku_to_hankaku_map = str.maketrans(zenkaku, hankaku)
    return text.translate(zenkaku_to_hankaku_map)

def remove_invisible_characters(text):
    return ''.join(c for c in text if not unicodedata.category(c).startswith('C'))


def normalize_text(text):
    text = unicodedata.normalize('NFKC', text)
    text = remove_invisible_characters(text)
    text = text.lower()
    text = re.sub(r'[\s\r\n]+', '', text)
    text = text.replace('㈱', '株式会社')
    text = re.sub(r'(株式会社|有限会社|合同会社|グループ|一般社団法人|一般財団法人|公益社団法人|公益財団法人|医療法人|学校法人|宗教法人|社会福祉法人|農業協同組合|漁業協同組合|生活協同組合|労働組合|特定非営利活動法人|独立行政法人|地方独立行政法人|特殊法人|特定目的会社|外国法人|㈱｜\（株\）|\(株\))', '', text)
    return text


def has_partial_match(g_name, h_name, length=4, threshold=3):
    if g_name == h_name:
        return True

    for i in range(len(g_name) - length + 1):
        substring = g_name[i:i + length]
        dist = distance(substring, h_name)
        if dist <= threshold:
            return True
    return False


def select_excel_file():
    root = Tk()
    root.withdraw()
    file_path = askopenfilename(title="エクセルファイルを選択", filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path


# エクセルファイルを選択
excel_file = select_excel_file()
if not excel_file:
    print("ファイルが選択されていません。終了します。")
    exit()

url = "https://www.invoice-kohyo.nta.go.jp/"
df = pd.read_excel(excel_file)

invoice_numbers = []
for index, value in enumerate(df['明細情報:フリー１(インボイス番号)']):
    if str(value) in ['N9999999999999', 'T9999999999999']:
        df.at[index, '企業名'] = "インボイス無し"
    else:
        number_only = re.sub(r'\D', '', str(value))
        invoice_numbers.append((index, number_only))

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--headless")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
driver.get(url)
time.sleep(2)

# インボイス番号を使って企業名とURLを取得
for index, number in invoice_numbers:
    try:
        if driver.current_url != url:
            driver.get(url)
            time.sleep(1)
            
        # インボイス番号を入力
        search_box = driver.find_element(By.ID, "regNo1")
        search_box.clear()
        search_box.send_keys(number)
        
        # 検索を実行
        search_button = driver.find_element(By.ID, "searchBtn")
        search_button.click()
        time.sleep(1)
        
        # 結果取得
        company_name_element = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "p.itemdata.sp_nmTsuushou_data"))
        )
        company_name = company_name_element.text.strip()

        result_url = driver.current_url  # 現在のURLを取得

        df.at[index, '企業名'] = company_name if company_name else "企業名未取得"
        df.at[index, 'I列'] = result_url  # URLをI列に格納

    except Exception as e:
        df.at[index, '企業名'] = "取得エラー"
        df.at[index, 'I列'] = ""

driver.quit()

# 現在の日付を取得してファイル名に埋め込む
current_date = datetime.now().strftime("%Y-%m-%d")
output_file = f"/Users/kazukiokada/Desktop/経費申請インボイス確認_{current_date}.xlsx"
df.to_excel(output_file, index=False)

# ハイパーリンク化
wb = load_workbook(output_file)
ws = wb.active

for row in range(2, ws.max_row + 1):
    url_value = ws.cell(row=row, column=9).value  # I列に対応
    if url_value and url_value.startswith("http"):
        ws.cell(row=row, column=9).hyperlink = url_value
        ws.cell(row=row, column=9).style = "Hyperlink"

# 黄色の塗りつぶしスタイル
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# データフレームの各行をループ
for index, row in df.iterrows():
    invoice_number = str(row['明細情報:フリー１(インボイス番号)'])
    g_column_name = normalize_text(str(row['明細情報:フリー２(支払先)']))
    h_column_name = normalize_text(str(row['企業名']))

    # 特定のインボイス番号をスキップ
    if invoice_number in ['N9999999999999', 'T9999999999999']:
        continue

    # 部分一致条件を満たさない場合
    if not has_partial_match(g_column_name, h_column_name, length=4, threshold=2):
        # 行全体を黄色でハイライト
        for col in range(1, ws.max_column + 1):  # 列を1から最終列までループ
            ws.cell(row=index + 2, column=col).fill = yellow_fill  # 行番号調整(index+2)
            
wb.save(output_file)
print(f"検索結果が保存されました: {output_file}")
